"""
LinkedIn Easy Apply Bot — Selenium automation engine.
No LinkedIn API required; drives a real Chrome browser session.

Excel file: linkedin_jobs.xlsx  (appended on every run, never overwritten)
"""

import re
import time
import random
import os
from datetime import datetime
from typing import Callable

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
)
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

try:
    from webdriver_manager.chrome import ChromeDriverManager
    USE_WDM = True
except ImportError:
    USE_WDM = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

EXCEL_FILE = "linkedin_jobs.xlsx"

DATE_FILTER_MAP = {
    "Past 24 hours": "r86400",
    "Past Week":     "r604800",
    "Past Month":    "r2592000",
    "Any Time":      "",
}

EXPERIENCE_FILTER_MAP = {
    "Internship":       "1",
    "Entry level":      "2",
    "Associate":        "3",
    "Mid-Senior level": "4",
    "Director":         "5",
    "Any":              "",
}

# All selectors tried in order; LinkedIn changes class names frequently
CARD_SELECTORS = [
    "li.jobs-search-results__list-item",
    "li[data-occludable-job-id]",
    "div.job-card-container",
    "div[data-job-id]",
    "li.scaffold-layout__list-item",
    "div.jobs-search-results__list > li",
    "ul > li.ember-view",
]

PAGE_READY_SELECTORS = [
    "div.jobs-search-results-list",
    "ul.jobs-search-results__list",
    "div.scaffold-layout__list",
    "div[class*='jobs-search-results']",
]

# ──────────────────────────────────────────────────────────────────────────────
# URL helpers
# ──────────────────────────────────────────────────────────────────────────────

def _build_search_url(role: str, location: str, date_posted: str, experience: str) -> str:
    # Wrap role in quotes → LinkedIn treats it as an exact job-title phrase,
    # not a general keyword search, giving far more targeted results.
    quoted = f'%22{role.replace(" ", "%20")}%22'
    params = [
        f"keywords={quoted}",
        f"location={location.replace(' ', '%20')}",
        "f_LF=f_AL",   # Easy Apply only
        "sortBy=DD",    # most recent first
    ]
    dc = DATE_FILTER_MAP.get(date_posted, "")
    if dc:
        params.append(f"f_TPR={dc}")
    ec = EXPERIENCE_FILTER_MAP.get(experience, "")
    if ec:
        params.append(f"f_E={ec}")
    return "https://www.linkedin.com/jobs/search/?" + "&".join(params)


def _job_id_from_url(url: str) -> str:
    """
    Extract the numeric LinkedIn job ID for deduplication.
    Handles two URL formats LinkedIn uses:
      - /jobs/view/1234567890/          (direct job page)
      - ?currentJobId=1234567890        (split-pane search view)
    Returns "" if no ID found — an empty string never matches anything.
    """
    if not url:
        return ""
    m = re.search(r"/jobs/view/(\d+)", url)
    if m:
        return m.group(1)
    m = re.search(r"currentJobId=(\d+)", url)
    if m:
        return m.group(1)
    return ""   # unknown — do NOT fall back to the URL itself


def _clean_job_url(raw_url: str) -> str:
    """
    Convert any LinkedIn job URL variant into a clean, canonical form:
    https://www.linkedin.com/jobs/view/JOBID/
    Falls back to the stripped URL if no job ID can be extracted.
    """
    job_id = _job_id_from_url(raw_url)
    if job_id:
        return f"https://www.linkedin.com/jobs/view/{job_id}/"
    return raw_url.split("?")[0]


# ──────────────────────────────────────────────────────────────────────────────
# Excel — persistent, append-only
# ──────────────────────────────────────────────────────────────────────────────

SHEET_META = {
    # Pipeline sheet — written at end of collection phase, status updated during apply phase
    "Collected Jobs": {
        "headers":   ["#", "Job Title", "Company", "URL", "Role Searched",
                      "Collected Date", "Collected Time", "Status", "Notes"],
        "hdr_color": "1A237E",
        "row_color": "E8EAF6",
    },
    "Applied": {
        "headers":     ["#", "Job Title", "Company", "URL", "Date", "Time", "Role Searched"],
        "hdr_color":   "1F4E79",
        "row_color":   "C6EFCE",
    },
    "Ignored": {
        "headers":     ["#", "Job Title", "Company", "URL", "Date", "Time", "Reason"],
        "hdr_color":   "7F6000",
        "row_color":   "FFEB9C",
    },
    "Failed": {
        "headers":     ["#", "Job Title", "Company", "URL", "Date", "Time", "Reason"],
        "hdr_color":   "9C0006",
        "row_color":   "FFC7CE",
    },
}

COL_WIDTHS = {
    "#":               5,
    "Job Title":       40,
    "Company":         28,
    "URL":             55,
    "Date":            14,
    "Collected Date":  14,
    "Time":            10,
    "Collected Time":  10,
    "Role Searched":   22,
    "Reason":          30,
    "Status":          16,
    "Notes":           35,
}


def _ensure_sheet(wb, name: str) -> "openpyxl.worksheet.worksheet.Worksheet":
    meta = SHEET_META[name]
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(meta["headers"])
        hf = PatternFill("solid", fgColor=meta["hdr_color"])
        for col_idx, h in enumerate(meta["headers"], 1):
            cell = ws.cell(1, col_idx)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = hf
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[cell.column_letter].width = COL_WIDTHS.get(h, 20)
    return wb[name]


def _load_or_create_workbook() -> "openpyxl.Workbook":
    if os.path.exists(EXCEL_FILE):
        try:
            return openpyxl.load_workbook(EXCEL_FILE)
        except Exception:
            pass
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb


def load_applied_job_ids(log: Callable) -> set:
    """
    Read linkedin_jobs.xlsx and return the set of already-applied LinkedIn job IDs.
    Called once at bot startup so we never re-apply to a previously applied job.
    """
    ids: set[str] = set()
    if not EXCEL_OK or not os.path.exists(EXCEL_FILE):
        return ids
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        if "Applied" in wb.sheetnames:
            ws = wb["Applied"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                url_cell = row[3] if len(row) > 3 else None
                if url_cell:
                    jid = _job_id_from_url(str(url_cell))
                    if jid:          # never add empty string
                        ids.add(jid)
        wb.close()
        if ids:
            log(f"[INFO] Loaded {len(ids)} previously applied job ID(s) from Excel — will skip them.")
        else:
            log("[INFO] No previous applications found in Excel — starting fresh.")
    except Exception as e:
        log(f"[WARN] Could not read existing Excel for dedup: {e}")
    return ids


def save_collected_jobs(jobs: list[dict], log: Callable) -> None:
    """
    Write the collected-jobs pipeline to the 'Collected Jobs' sheet.
    Called at the end of Phase 1.  Existing rows are preserved; new ones appended.
    Each job dict must have: title, company, url, role, collected_at, status, notes.
    """
    if not EXCEL_OK:
        log("[WARN] openpyxl not installed — cannot save collected jobs.")
        return
    if not jobs:
        return
    try:
        wb  = _load_or_create_workbook()
        ws  = _ensure_sheet(wb, "Collected Jobs")
        rf  = PatternFill("solid", fgColor=SHEET_META["Collected Jobs"]["row_color"])
        start = ws.max_row   # 1 = header row only

        for offset, j in enumerate(jobs):
            row_num   = start + offset + 1
            row_index = row_num - 1
            dt        = datetime.strptime(j["collected_at"], "%Y-%m-%d %H:%M:%S")
            ws.append([
                row_index,
                j["title"],
                j["company"],
                j["url"],
                j.get("role", ""),
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                j.get("status", "Pending"),
                j.get("notes", ""),
            ])
            for col in range(1, 10):
                ws.cell(row_num, col).fill = rf

        wb.save(EXCEL_FILE)
        log(f"[INFO] Collected {len(jobs)} jobs saved → {os.path.abspath(EXCEL_FILE)}")
    except Exception as e:
        log(f"[ERROR] Could not save collected jobs: {e}")


def update_collected_status(excel_row: int, status: str, notes: str, log: Callable) -> None:
    """
    Update a single row in 'Collected Jobs' sheet with a new status and notes.
    excel_row is 1-based (2 = first data row).
    Called after each apply attempt in Phase 2 so progress is visible in real time.
    """
    if not EXCEL_OK or not os.path.exists(EXCEL_FILE):
        return
    # Status = col 8, Notes = col 9
    STATUS_COL = 8
    NOTES_COL  = 9
    STATUS_COLORS = {
        "Applied":        "C6EFCE",
        "Failed":         "FFC7CE",
        "Skipped":        "FFEB9C",
        "No Easy Apply":  "FCE4EC",
        "Already Applied":"E0E0E0",
    }
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Collected Jobs"]
        ws.cell(excel_row, STATUS_COL).value = status
        ws.cell(excel_row, NOTES_COL).value  = notes
        color = STATUS_COLORS.get(status, "E8EAF6")
        fill  = PatternFill("solid", fgColor=color)
        for col in range(1, 10):
            ws.cell(excel_row, col).fill = fill
        wb.save(EXCEL_FILE)
    except Exception as e:
        log(f"[WARN] Could not update row {excel_row} status: {e}")


def update_excel(
    applied: list[dict],
    ignored: list[dict],
    failed:  list[dict],
    log: Callable,
) -> None:
    if not EXCEL_OK:
        log("[WARN] openpyxl not installed — skipping Excel export. Run: pip install openpyxl")
        return
    if not (applied or ignored or failed):
        log("[INFO] Nothing new to save to Excel.")
        return

    try:
        wb = _load_or_create_workbook()

        for sheet_name, jobs in [("Applied", applied), ("Ignored", ignored), ("Failed", failed)]:
            if not jobs:
                continue
            ws    = _ensure_sheet(wb, sheet_name)
            start = ws.max_row      # 1 = header only → next data row = 2
            rf    = PatternFill("solid", fgColor=SHEET_META[sheet_name]["row_color"])

            for offset, j in enumerate(jobs):
                row_num    = start + offset + 1
                row_index  = row_num - 1          # sequential # ignoring header
                dt         = datetime.strptime(j["timestamp"], "%Y-%m-%d %H:%M:%S")
                date_str   = dt.strftime("%Y-%m-%d")
                time_str   = dt.strftime("%H:%M:%S")

                if sheet_name == "Applied":
                    values = [row_index, j["title"], j["company"], j["url"],
                              date_str, time_str, j.get("role", "")]
                else:
                    values = [row_index, j["title"], j["company"], j["url"],
                              date_str, time_str, j.get("reason", "")]

                ws.append(values)
                for col_idx in range(1, len(values) + 1):
                    ws.cell(row_num, col_idx).fill = rf

        wb.save(EXCEL_FILE)
        path = os.path.abspath(EXCEL_FILE)
        log(f"[INFO] Excel updated: {path}  "
            f"(+{len(applied)} applied, +{len(ignored)} ignored, +{len(failed)} failed)")

    except Exception as e:
        log(f"[ERROR] Could not update Excel: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Bot
# ──────────────────────────────────────────────────────────────────────────────

class LinkedInBot:

    def __init__(
        self,
        config: dict,
        log: Callable[[str], None],
        collect_callback: Callable[[int], None] | None = None,
        apply_callback:   Callable[[int, int], None] | None = None,
    ):
        self.config           = config
        self.log              = log
        self.collect_callback = collect_callback  # (collected_count)
        self.apply_callback   = apply_callback    # (applied_count, total_jobs)
        self.driver: webdriver.Chrome | None = None
        self.wait:   WebDriverWait    | None = None
        self._stop_flag       = False
        self.applied_count    = 0
        self._total_to_apply  = 0   # set at start of Phase 2

        # Per-run job lists
        self.applied_jobs: list[dict] = []
        self.ignored_jobs: list[dict] = []
        self.failed_jobs:  list[dict] = []

        # Loaded from Excel before run starts
        self._applied_ids: set[str] = set()

    # ─────────────────────────────────── control

    def request_stop(self):
        self._stop_flag = True

    def _should_stop(self) -> bool:
        return self._stop_flag

    # ─────────────────────────────────── delays

    def _delay(self, lo: float | None = None, hi: float | None = None):
        lo = lo if lo is not None else float(self.config.get("min_delay", 3))
        hi = hi if hi is not None else float(self.config.get("max_delay", 7))
        time.sleep(random.uniform(lo, hi))

    def _short(self):
        time.sleep(random.uniform(0.7, 1.5))

    # ─────────────────────────────────── browser setup

    def _setup_driver(self):
        opts = Options()
        if self.config.get("headless", False):
            opts.add_argument("--headless=new")

        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--start-maximized")
        opts.add_argument("--window-size=1440,900")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )

        if USE_WDM:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=opts)
        else:
            self.driver = webdriver.Chrome(options=opts)

        self.driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
        )
        self.wait = WebDriverWait(self.driver, 15)
        self.log("[INFO] Browser launched.")

    # ─────────────────────────────────── login

    def _login(self) -> bool:
        self.log("[INFO] Opening LinkedIn login page...")
        self.driver.get("https://www.linkedin.com/login")
        self._delay(2, 4)

        try:
            email_f = self.wait.until(EC.presence_of_element_located((By.ID, "username")))
            self._type_human(email_f, self.config["email"])
            pw = self.driver.find_element(By.ID, "password")
            self._type_human(pw, self.config["password"])
            self._short()
            pw.send_keys(Keys.RETURN)
            self._delay(4, 6)
        except TimeoutException:
            self.log("[ERROR] Login page did not load.")
            return False

        url = self.driver.current_url
        if any(k in url for k in ("feed", "mynetwork", "jobs", "home")):
            self.log("[INFO] Login successful.")
            return True

        if any(k in url for k in ("checkpoint", "challenge", "security", "verification")):
            self.log("[WARN] Security check required — complete it in the browser (90 s).")
            deadline = time.time() + 90
            while time.time() < deadline:
                time.sleep(3)
                if any(k in self.driver.current_url for k in ("feed", "mynetwork", "jobs")):
                    self.log("[INFO] Security check passed.")
                    return True
            self.log("[ERROR] Security check timed out.")
            return False

        self.log("[ERROR] Login failed — unexpected URL: " + url)
        return False

    def _type_human(self, element, text: str):
        element.clear()
        for ch in text:
            element.send_keys(ch)
            time.sleep(random.uniform(0.04, 0.13))

    # ─────────────────────────────────── search

    def _search(self, role: str):
        url = _build_search_url(
            role,
            self.config.get("location", "United States"),
            self.config.get("date_posted", "Any Time"),
            self.config.get("experience_level", "Any"),
        )
        self.log(f"[INFO] Searching: {role}")
        self.driver.get(url)
        self._delay(4, 6)

    # ─────────────────────────────────── card discovery

    def _wait_for_page_ready(self):
        combined = ", ".join(PAGE_READY_SELECTORS)
        try:
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, combined))
            )
        except TimeoutException:
            self.log("[WARN] Page ready timeout — proceeding anyway.")

    def _get_job_cards(self) -> list:
        self._wait_for_page_ready()
        time.sleep(1.5)

        for sel in CARD_SELECTORS:
            cards   = self.driver.find_elements(By.CSS_SELECTOR, sel)
            visible = [c for c in cards if c.is_displayed()]
            if len(visible) > 1:
                self.log(f"[DEBUG] {len(visible)} cards  ({sel})")
                return visible

        self.log(f"[DEBUG] No cards found — page title: '{self.driver.title}'")
        self.log(f"[DEBUG] URL: {self.driver.current_url}")
        return []

    # ─────────────────────────────────── card data extraction

    def _get_title(self, card) -> str:
        for sel in [
            "a.job-card-list__title--link",
            "a.job-card-list__title",
            "a.job-card-container__link",
            "strong",
            "a[href*='/jobs/view/']",
        ]:
            try:
                t = card.find_element(By.CSS_SELECTOR, sel).text.strip()
                if t:
                    return t
            except NoSuchElementException:
                continue
        return "Unknown Title"

    def _get_company(self, card) -> str:
        for sel in [
            ".job-card-container__company-name",
            ".job-card-container__primary-description",
            "a.job-card-container__company-name",
            ".artdeco-entity-lockup__subtitle span",
            ".job-card-list__company-name",
            "span.job-card-container__company-name",
        ]:
            try:
                t = card.find_element(By.CSS_SELECTOR, sel).text.strip()
                if t:
                    return t
            except NoSuchElementException:
                continue
        return "Unknown Company"

    def _get_url(self, card) -> str:
        """
        Get the job URL from the card's anchor tag href.
        Card hrefs always contain /jobs/view/JOBID/ so we can extract the ID
        BEFORE clicking — avoiding the split-pane currentJobId problem entirely.
        Returns "" if not found (caller must handle).
        """
        for sel in [
            "a.job-card-list__title--link",
            "a.job-card-list__title",
            "a.job-card-container__link",
            "a[href*='/jobs/view/']",
        ]:
            try:
                href = card.find_element(By.CSS_SELECTOR, sel).get_attribute("href") or ""
                if "/jobs/view/" in href:
                    return _clean_job_url(href)
            except NoSuchElementException:
                continue
        # Try the data attribute on the card itself
        try:
            jid = card.get_attribute("data-occludable-job-id") or card.get_attribute("data-job-id")
            if jid:
                return f"https://www.linkedin.com/jobs/view/{jid}/"
        except Exception:
            pass
        return ""

    def _get_detail_text(self) -> str:
        """Read the job detail panel text after clicking a card."""
        for sel in [
            "div.jobs-description",
            "div.jobs-unified-top-card",
            "div.job-details-jobs-unified-top-card__job-insight",
            "div.jobs-details",
        ]:
            try:
                return self.driver.find_element(By.CSS_SELECTOR, sel).text.lower()
            except NoSuchElementException:
                continue
        return ""

    # ─────────────────────────────────── filtering logic

    def _check_ignore(self, title: str) -> tuple[bool, str]:
        """Returns (should_skip, matched_word)."""
        lower = title.lower()
        for word in self.config.get("ignore_words", []):
            if word.strip().lower() in lower:
                return True, word
        return False, ""

    def _check_role_match(self, title: str) -> bool:
        """
        If strict_role_match is on, skip any job whose title contains none
        of the keywords derived from the user's job_roles list.
        Splits each role into words >3 chars and checks any of them.
        """
        if not self.config.get("strict_role_match", True):
            return True
        title_lower = title.lower()
        for role in self.config.get("job_roles", []):
            keywords = [w for w in role.lower().split() if len(w) > 3]
            if any(kw in title_lower for kw in keywords):
                return True
        return False

    def _check_job_type(self, title: str, detail_text: str) -> tuple[bool, str]:
        """
        If job_type_keywords list is non-empty, at least one keyword must appear
        in the title or job detail text.
        Returns (passes, matched_keyword_or_reason).
        """
        required = [k.strip() for k in self.config.get("job_type_keywords", []) if k.strip()]
        if not required:
            return True, ""
        combined = (title + " " + detail_text).lower()
        for kw in required:
            if kw.lower() in combined:
                return True, kw
        return False, f"none of {required} found"

    # ─────────────────────────────────── apply flow

    def _click_card(self, card) -> bool:
        try:
            self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", card)
            self._short()
            card.click()
            self._delay(2, 3)
            return True
        except Exception as e:
            self.log(f"[WARN] Could not click card: {e}")
            return False

    def _click_easy_apply(self) -> bool:
        for sel in [
            "button.jobs-apply-button",
            "button[aria-label*='Easy Apply']",
            "button[aria-label*='easy apply']",
            ".jobs-s-apply button",
        ]:
            try:
                btn = WebDriverWait(self.driver, 6).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
                )
                label = (btn.text or btn.get_attribute("aria-label") or "").lower()
                if "easy apply" in label:
                    btn.click()
                    self._delay(1.5, 3)
                    return True
            except Exception:
                continue
        return False

    def _find_btn(self, *fragments: str):
        for btn in self.driver.find_elements(By.TAG_NAME, "button"):
            if not btn.is_displayed():
                continue
            label = (btn.get_attribute("aria-label") or "").lower()
            text  = (btn.text or "").lower()
            for f in fragments:
                if f.lower() in label or f.lower() in text:
                    return btn
        return None

    def _click_next_or_review(self) -> bool:
        btn = self._find_btn(
            "continue to next step", "next step", "next",
            "review your application", "review",
        )
        if btn and btn.is_enabled():
            try:
                btn.click()
                self._delay(1.5, 2.5)
                return True
            except Exception:
                pass
        return False

    def _click_submit(self) -> bool:
        btn = self._find_btn("submit application", "submit")
        if btn and btn.is_enabled():
            try:
                btn.click()
                self._delay(2, 3.5)
                return True
            except Exception:
                pass
        return False

    def _close_modal(self):
        for sel in [
            "button[aria-label='Dismiss']",
            "button[aria-label='Close']",
            ".artdeco-modal__dismiss",
            "button[data-test-modal-close-btn]",
        ]:
            try:
                btn = WebDriverWait(self.driver, 4).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, sel))
                )
                btn.click()
                self._short()
                return
            except Exception:
                continue

    def _discard(self):
        btn = self._find_btn("discard", "discard application")
        if btn:
            try:
                btn.click()
                self._short()
            except Exception:
                pass

    # ─────────────────────────────────── form auto-fill

    def _get_label_text(self, element) -> str:
        """Return the question/label text associated with a form element."""
        v = (element.get_attribute("aria-label") or "").strip()
        if v:
            return v
        elem_id = element.get_attribute("id") or ""
        if elem_id:
            try:
                return self.driver.find_element(
                    By.CSS_SELECTOR, f'label[for="{elem_id}"]'
                ).text.strip()
            except Exception:
                pass
        try:
            return element.find_element(
                By.XPATH, "..//label"
            ).text.strip()
        except Exception:
            pass
        return (element.get_attribute("placeholder") or "").strip()

    def _map_answer(self, label: str, answers: dict) -> str:
        """Map a question label to a user-configured answer string."""
        lbl = label.lower()
        if any(w in lbl for w in ("phone", "mobile", "telephone")):
            return answers.get("phone", "")
        if any(w in lbl for w in ("year", "experience", "how long", "how many")):
            return str(answers.get("years_experience", "2"))
        if any(w in lbl for w in ("annual salary", "salary", "compensation", "ctc")):
            return str(answers.get("expected_salary", ""))
        if any(w in lbl for w in ("rate", "hourly", "per hour")):
            return str(answers.get("expected_rate", ""))
        if any(w in lbl for w in ("city", "current city")):
            return answers.get("city", "")
        if any(w in lbl for w in ("state", "province")):
            return answers.get("state", "")
        if any(w in lbl for w in ("zip", "postal")):
            return answers.get("zip_code", "")
        if "country" in lbl:
            return answers.get("country", "United States")
        if any(w in lbl for w in ("linkedin url", "linkedin profile", "linkedin.com")):
            return answers.get("linkedin_url", "")
        if any(w in lbl for w in ("github", "portfolio", "website", "personal site")):
            return answers.get("portfolio_url", "")
        return ""

    def _fill_text_inputs(self, answers: dict):
        try:
            inputs = self.driver.find_elements(
                By.CSS_SELECTOR,
                "input[type='text'], input[type='number'], input[type='tel'], input[type='email']",
            )
            for inp in inputs:
                if not inp.is_displayed() or not inp.is_enabled():
                    continue
                current = (inp.get_attribute("value") or "").strip()
                if current and current not in ("0",):
                    continue  # already filled
                label = self._get_label_text(inp)
                value = self._map_answer(label, answers)
                if value:
                    inp.clear()
                    inp.send_keys(value)
                    self._short()
        except Exception:
            pass

    def _select_best_option(self, sel_obj, target: str):
        """Pick the select option whose text best matches target."""
        if not target:
            return
        t = target.lower()
        try:
            for opt in sel_obj.options:
                if opt.text.strip().lower() == t:
                    sel_obj.select_by_visible_text(opt.text)
                    return
            for opt in sel_obj.options:
                if t in opt.text.strip().lower():
                    sel_obj.select_by_visible_text(opt.text)
                    return
            for opt in sel_obj.options:
                ot = opt.text.strip().lower()
                if ot and ot in t:
                    sel_obj.select_by_visible_text(opt.text)
                    return
        except Exception:
            pass

    def _fill_selects(self, answers: dict):
        from selenium.webdriver.support.ui import Select as SeleniumSelect
        try:
            for sel_el in self.driver.find_elements(By.TAG_NAME, "select"):
                if not sel_el.is_displayed() or not sel_el.is_enabled():
                    continue
                s = SeleniumSelect(sel_el)
                current = s.first_selected_option.text.strip().lower()
                if current not in ("select an option", "please select", "", "select", "-",
                                   "-- select --", "choose an option"):
                    continue  # already has a selection
                lbl = self._get_label_text(sel_el).lower()
                if any(w in lbl for w in ("education", "degree", "qualification")):
                    self._select_best_option(s, answers.get("education_level", "Bachelor's Degree"))
                elif "country" in lbl:
                    self._select_best_option(s, answers.get("country", "United States"))
                elif any(w in lbl for w in ("state", "province")):
                    self._select_best_option(s, answers.get("state", ""))
                elif any(w in lbl for w in ("experience level", "seniority")):
                    self._select_best_option(s, answers.get("exp_level_label", "Entry level"))
                elif any(w in lbl for w in ("employment type", "job type", "work type")):
                    self._select_best_option(s, answers.get("work_type", "Full-time"))
                else:
                    # Default: pick first non-empty option
                    try:
                        for opt in s.options[1:]:
                            if opt.text.strip():
                                s.select_by_visible_text(opt.text)
                                break
                    except Exception:
                        pass
        except Exception:
            pass

    def _fill_radio_groups(self, answers: dict):
        try:
            for fs in self.driver.find_elements(By.TAG_NAME, "fieldset"):
                if not fs.is_displayed():
                    continue
                radios = fs.find_elements(By.CSS_SELECTOR, "input[type='radio']")
                if not radios or any(r.is_selected() for r in radios):
                    continue
                question = ""
                try:
                    question = fs.find_element(By.TAG_NAME, "legend").text.lower()
                except Exception:
                    pass
                if any(w in question for w in ("authorized", "eligible", "legally", "work in")):
                    target = answers.get("authorized_to_work", "Yes")
                elif any(w in question for w in ("sponsor", "sponsorship", "visa")):
                    target = answers.get("require_sponsorship", "No")
                elif any(w in question for w in ("relocat", "willing to move")):
                    target = answers.get("willing_to_relocate", "No")
                elif any(w in question for w in ("remote", "hybrid", "on-site", "work mode", "work location")):
                    target = answers.get("work_preference", "Remote")
                else:
                    target = answers.get("default_yes_no", "Yes")
                clicked = False
                for radio in radios:
                    try:
                        rid = radio.get_attribute("id") or ""
                        rlabel = ""
                        if rid:
                            try:
                                rlabel = self.driver.find_element(
                                    By.CSS_SELECTOR, f'label[for="{rid}"]'
                                ).text.strip()
                            except Exception:
                                pass
                        if not rlabel:
                            rlabel = radio.get_attribute("value") or ""
                        if target.lower() in rlabel.lower():
                            self.driver.execute_script("arguments[0].click();", radio)
                            clicked = True
                            break
                    except Exception:
                        continue
                if not clicked:
                    try:
                        self.driver.execute_script("arguments[0].click();", radios[0])
                    except Exception:
                        pass
        except Exception:
            pass

    def _fill_textareas(self, answers: dict):
        cover = answers.get("cover_letter", "").strip()
        if not cover:
            return
        try:
            for ta in self.driver.find_elements(By.TAG_NAME, "textarea"):
                if not ta.is_displayed() or not ta.is_enabled():
                    continue
                if (ta.get_attribute("value") or ta.text or "").strip():
                    continue
                lbl = self._get_label_text(ta).lower()
                if any(w in lbl for w in ("cover", "letter", "summary", "about", "message", "additional")):
                    ta.clear()
                    ta.send_keys(cover)
                    self._short()
        except Exception:
            pass

    def _fill_form_fields(self):
        """Fill any empty required fields on the current form step."""
        answers = self.config.get("form_answers", {})
        try:
            self._fill_text_inputs(answers)
            self._fill_selects(answers)
            self._fill_radio_groups(answers)
            self._fill_textareas(answers)
        except Exception as e:
            self.log(f"[WARN] Form fill error (non-fatal): {e}")

    def _handle_form(self) -> tuple[bool, str]:
        for step in range(15):
            if self._should_stop():
                self._close_modal()
                self._discard()
                return False, "stopped by user"
            self._delay(1, 2)
            # Fill any empty fields on this step before navigating
            self._fill_form_fields()
            self._short()
            if self._click_submit():
                self._delay(1, 2)
                self._close_modal()
                return True, ""
            if self._click_next_or_review():
                continue
            self._close_modal()
            self._discard()
            return False, f"no nav button at step {step + 1}"
        self._close_modal()
        self._discard()
        return False, "exceeded max form steps"

    # ─────────────────────────────────── pagination

    def _next_page(self, current: int) -> bool:
        try:
            btn = self.driver.find_element(
                By.CSS_SELECTOR, f'button[aria-label="Page {current + 1}"]'
            )
            btn.click()
            self._delay(3, 5)
            return True
        except NoSuchElementException:
            return False

    def _scroll_list(self):
        for sel in [
            "div.jobs-search-results-list",
            "ul.jobs-search-results__list",
            "div.scaffold-layout__list",
        ]:
            try:
                panel = self.driver.find_element(By.CSS_SELECTOR, sel)
                self.driver.execute_script(
                    "arguments[0].scrollTop = arguments[0].scrollHeight", panel
                )
                self._delay(1.5, 2.5)
                return
            except Exception:
                continue
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight / 3);")
        self._delay(1.5, 2)

    # ─────────────────────────────────── Phase 1 — collect

    def _collect_phase(self) -> list[dict]:
        """
        Fast collection pass — only reads job card metadata (title, company, URL).
        Never clicks into a job detail page.
        Applies: already-applied check, ignore list, strict role match.
        Job-type keyword check is deferred to Phase 2 (needs the detail page).
        Returns a list of job dicts with status='Pending'.
        """
        collected: list[dict] = []
        seen_ids:  set[str]   = set()
        roles = self.config.get("job_roles", [])

        for role in roles:
            if self._should_stop():
                break

            self._search(role)
            page = 1

            while not self._should_stop():
                self._scroll_list()
                cards = self._get_job_cards()

                if not cards:
                    self.log(f"[INFO] No more cards for '{role}' (page {page}).")
                    break

                seen_titles: set[str] = set()

                for card in cards:
                    if self._should_stop():
                        break
                    try:
                        title   = self._get_title(card)
                        company = self._get_company(card)
                        url     = self._get_url(card)
                        job_id  = _job_id_from_url(url)

                        if title in seen_titles or not url:
                            continue
                        seen_titles.add(title)

                        if job_id and job_id in self._applied_ids:
                            self.log(f"[SKIP] Already applied: {title}")
                            continue
                        if job_id and job_id in seen_ids:
                            continue
                        seen_ids.add(job_id or url)

                        skip, word = self._check_ignore(title)
                        if skip:
                            self.log(f"[SKIP] Ignored ('{word}'): {title}")
                            continue

                        if not self._check_role_match(title):
                            self.log(f"[SKIP] Off-role: {title}")
                            continue

                        collected.append({
                            "title":        title,
                            "company":      company,
                            "url":          url,
                            "job_id":       job_id,
                            "role":         role,
                            "status":       "Pending",
                            "notes":        "",
                            "collected_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        })
                        self.log(f"[COLLECT] #{len(collected)}  {title}  @  {company}")

                        if self.collect_callback:
                            self.collect_callback(len(collected))

                    except StaleElementReferenceException:
                        break
                    except Exception as exc:
                        self.log(f"[WARN] Card read error: {exc}")
                        continue

                if not self._next_page(page):
                    break
                page += 1

        return collected

    # ─────────────────────────────────── Phase 2 — apply

    def _apply_phase(self, jobs: list[dict]) -> None:
        """
        Application pass — navigates to each collected job URL directly.
        No search-page scraping. Updates Excel row status after every attempt.
        """
        max_apps  = int(self.config.get("max_applications", 50))
        self._total_to_apply = len(jobs)
        # jobs[0] is at Excel row 2 (header = row 1)
        base_row  = 2

        for i, job in enumerate(jobs):
            if self._should_stop() or self.applied_count >= max_apps:
                break

            excel_row = base_row + i
            title     = job["title"]
            company   = job["company"]
            url       = job["url"]
            role      = job.get("role", "")
            now       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            total     = len(jobs)

            self.log(f"\n[INFO] ── Applying {i+1}/{total}: {title} @ {company}")

            # Navigate directly to the job page — no re-scraping
            try:
                self.driver.get(url)
                self._delay(2, 4)
            except Exception as e:
                self.log(f"[WARN] Could not load job page: {e}")
                job["status"] = "Failed"
                job["notes"]  = "page load error"
                update_collected_status(excel_row, "Failed", "page load error", self.log)
                continue

            # Resolve URL in case of redirect (LinkedIn might use currentJobId)
            final_url = _clean_job_url(self.driver.current_url)
            job_id    = _job_id_from_url(final_url)
            if job_id and job_id in self._applied_ids:
                self.log(f"[SKIP] Already applied: {title}")
                job["status"] = "Already Applied"
                update_collected_status(excel_row, "Already Applied", "", self.log)
                continue

            # Job-type keyword check (needs page content)
            detail_text       = self._get_detail_text()
            type_ok, type_why = self._check_job_type(title, detail_text)
            if not type_ok:
                self.log(f"[SKIP] Job type mismatch ({type_why}): {title}")
                job["status"] = "Skipped"
                job["notes"]  = type_why
                update_collected_status(excel_row, "Skipped", type_why, self.log)
                self.ignored_jobs.append({
                    "title": title, "company": company, "url": final_url,
                    "timestamp": now, "reason": f"job type: {type_why}", "role": role,
                })
                continue

            if not self._click_easy_apply():
                self.log(f"[SKIP] No Easy Apply button: {title}")
                job["status"] = "No Easy Apply"
                update_collected_status(excel_row, "No Easy Apply", "", self.log)
                self.failed_jobs.append({
                    "title": title, "company": company, "url": final_url,
                    "timestamp": now, "reason": "no Easy Apply button", "role": role,
                })
                continue

            success, fail_reason = self._handle_form()

            if success:
                self.applied_count += 1
                self._applied_ids.add(job_id)
                job["status"] = "Applied"
                update_collected_status(excel_row, "Applied", "", self.log)
                self.applied_jobs.append({
                    "title": title, "company": company, "url": final_url,
                    "timestamp": now, "reason": "", "role": role,
                })
                if self.apply_callback:
                    self.apply_callback(self.applied_count, self._total_to_apply)
                self.log(f"[SUCCESS] Applied ({self.applied_count}/{max_apps}): {title} @ {company}")
            else:
                job["status"] = "Failed"
                job["notes"]  = fail_reason
                update_collected_status(excel_row, "Failed", fail_reason, self.log)
                self.failed_jobs.append({
                    "title": title, "company": company, "url": final_url,
                    "timestamp": now, "reason": fail_reason, "role": role,
                })
                self.log(f"[WARN] Failed ({fail_reason}): {title}")

            self._delay()

    # ─────────────────────────────────── main run

    def run(self, collect_only: bool = False):
        """
        Two-phase execution:
          Phase 1 — Collect: scrape all matching job cards (fast, no detail clicks).
                              Save to 'Collected Jobs' sheet in linkedin_jobs.xlsx.
          Phase 2 — Apply:   navigate to each URL directly and apply.
                              Update status in Excel row-by-row as we go.
        Pass collect_only=True to stop after Phase 1 (no applications made).
        """
        try:
            self._setup_driver()
            if not self._login():
                return

            self._applied_ids = load_applied_job_ids(self.log)

            # ── Phase 1 ──────────────────────────────────────────
            self.log("\n[INFO] ═══ PHASE 1 — Collecting jobs ═══")
            collected = self._collect_phase()

            self.log(f"[INFO] Collected {len(collected)} jobs matching your filters.")
            save_collected_jobs(collected, self.log)

            if collect_only or self._should_stop() or not collected:
                if collect_only:
                    self.log("[INFO] Collect-only mode — skipping application phase.")
                elif not collected:
                    self.log("[INFO] Nothing to apply to.")
                return

            # ── Phase 2 ──────────────────────────────────────────
            self.log(f"\n[INFO] ═══ PHASE 2 — Applying to {len(collected)} jobs ═══")
            self._apply_phase(collected)

            self.log(
                f"\n[DONE] Applied: {len(self.applied_jobs)}  |  "
                f"Skipped/Ignored: {len(self.ignored_jobs)}  |  "
                f"Failed: {len(self.failed_jobs)}"
            )
            update_excel(self.applied_jobs, self.ignored_jobs, self.failed_jobs, self.log)

        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                self.log("[INFO] Browser closed.")


# ──────────────────────────────────────────────────────────────────────────────
# Feed Excel — separate file, same append-only pattern
# ──────────────────────────────────────────────────────────────────────────────

FEED_EXCEL_FILE = "linkedin_feed_posts.xlsx"

FEED_SHEET_META = {
    "Feed Posts": {
        "headers":   ["#", "Author", "Author Profile URL", "Post URL",
                      "Post Preview", "Keywords Found", "Date", "Time"],
        "hdr_color": "4A148C",
        "row_color": "EDE7F6",
    }
}

FEED_COL_WIDTHS = {
    "#":                 5,
    "Author":           25,
    "Author Profile URL": 45,
    "Post URL":         55,
    "Post Preview":     60,
    "Keywords Found":   25,
    "Date":             14,
    "Time":             10,
}


def update_feed_excel(posts: list[dict], log: Callable) -> None:
    if not EXCEL_OK:
        log("[WARN] openpyxl not installed — skipping Excel. Run: pip install openpyxl")
        return
    if not posts:
        log("[INFO] No new feed posts to save.")
        return

    try:
        if os.path.exists(FEED_EXCEL_FILE):
            wb = openpyxl.load_workbook(FEED_EXCEL_FILE)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        sheet_name = "Feed Posts"
        meta = FEED_SHEET_META[sheet_name]

        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(meta["headers"])
            hf = PatternFill("solid", fgColor=meta["hdr_color"])
            for col_idx, h in enumerate(meta["headers"], 1):
                cell = ws.cell(1, col_idx)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = hf
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[cell.column_letter].width = FEED_COL_WIDTHS.get(h, 20)
        else:
            ws = wb[sheet_name]

        rf    = PatternFill("solid", fgColor=meta["row_color"])
        start = ws.max_row

        for offset, p in enumerate(posts):
            row_num   = start + offset + 1
            row_index = row_num - 1
            dt        = datetime.strptime(p["timestamp"], "%Y-%m-%d %H:%M:%S")
            ws.append([
                row_index,
                p.get("author", ""),
                p.get("author_url", ""),
                p.get("post_url", ""),
                p.get("preview", "")[:300],   # cap at 300 chars
                p.get("keywords_found", ""),
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
            ])
            for col_idx in range(1, 9):
                ws.cell(row_num, col_idx).fill = rf

        wb.save(FEED_EXCEL_FILE)
        log(f"[INFO] Feed Excel updated: {os.path.abspath(FEED_EXCEL_FILE)}  (+{len(posts)} posts)")

    except Exception as e:
        log(f"[ERROR] Could not update feed Excel: {e}")


# ──────────────────────────────────────────────────────────────────────────────
# Feed Scanner
# ──────────────────────────────────────────────────────────────────────────────

# ── Post result containers (tried in order for both search and feed pages) ──
POST_CONTAINER_SELECTORS = [
    # LinkedIn content/post search results page
    "li.reusable-search__result-container",
    "div.search-results__list > li",
    # Feed page fallback
    "div.feed-shared-update-v2",
    "div[data-urn*='activity']",
    "li.occludable-update",
]

# ── Post body text ──
POST_TEXT_SELECTORS = [
    ".update-components-text span.break-words",
    ".update-components-text",
    "div.feed-shared-update-v2__description span.break-words",
    "div.feed-shared-update-v2__description",
    ".attributed-text-segment-list__content",
    "span.break-words",
    # content search snippet
    ".reusable-search-simple-insight__text",
    ".entity-result__simple-insight-text",
    "p[class*='insight']",
]

# ── Author name ──
AUTHOR_NAME_SELECTORS = [
    ".update-components-actor__name span[aria-hidden='true']",
    ".update-components-actor__name",
    ".feed-shared-actor__name",
    # content search author
    ".entity-result__title-text a",
    "span.entity-result__title-text",
    "a.app-aware-link[href*='/in/']",
]

# ── Author profile URL ──
AUTHOR_URL_SELECTORS = [
    "a.update-components-actor__container",
    "a.feed-shared-actor__container-link",
    # content search
    ".entity-result__title-text a[href*='/in/']",
    "a[href*='/in/'][class*='result']",
    "a[href*='/in/']",
]


class FeedScanner:
    """
    Logs in to LinkedIn, scrolls the feed, and saves any post that contains
    one or more of the configured keywords to linkedin_feed_posts.xlsx.
    Runs independently from LinkedInBot — opens its own Chrome window.
    """

    def __init__(
        self,
        config: dict,
        log: Callable[[str], None],
        count_callback: Callable[[int], None] | None = None,
    ):
        self.config         = config
        self.log            = log
        self.count_callback = count_callback
        self.driver: webdriver.Chrome | None = None
        self.wait:   WebDriverWait    | None = None
        self._stop_flag  = False
        self.found_posts: list[dict] = []
        self._seen_urns:  set[str]   = set()   # deduplicate within a session

    def request_stop(self):
        self._stop_flag = True

    def _should_stop(self) -> bool:
        return self._stop_flag

    # ─────────────────────────────────── delays

    def _delay(self, lo: float = 2.0, hi: float = 4.0):
        time.sleep(random.uniform(lo, hi))

    def _short(self):
        time.sleep(random.uniform(0.6, 1.2))

    # ─────────────────────────────────── browser (same setup as LinkedInBot)

    def _setup_driver(self):
        opts = Options()
        if self.config.get("headless", False):
            opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument("--start-maximized")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        if USE_WDM:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=opts)
        else:
            self.driver = webdriver.Chrome(options=opts)

        self.driver.execute_cdp_cmd(
            "Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
        )
        self.wait = WebDriverWait(self.driver, 15)
        self.log("[INFO] Feed scanner browser launched.")

    # ─────────────────────────────────── login (same as LinkedInBot)

    def _login(self) -> bool:
        self.log("[INFO] Logging in for feed scan...")
        self.driver.get("https://www.linkedin.com/login")
        self._delay(2, 4)
        try:
            ef = self.wait.until(EC.presence_of_element_located((By.ID, "username")))
            for ch in self.config["email"]:
                ef.send_keys(ch)
                time.sleep(random.uniform(0.04, 0.12))
            pw = self.driver.find_element(By.ID, "password")
            for ch in self.config["password"]:
                pw.send_keys(ch)
                time.sleep(random.uniform(0.04, 0.12))
            self._short()
            pw.send_keys(Keys.RETURN)
            self._delay(4, 6)
        except TimeoutException:
            self.log("[ERROR] Login page did not load.")
            return False

        url = self.driver.current_url
        if any(k in url for k in ("feed", "mynetwork", "jobs", "home")):
            self.log("[INFO] Login successful.")
            return True

        if any(k in url for k in ("checkpoint", "challenge", "security")):
            self.log("[WARN] Security check — complete it in the browser (90 s).")
            deadline = time.time() + 90
            while time.time() < deadline:
                time.sleep(3)
                if any(k in self.driver.current_url for k in ("feed", "mynetwork", "jobs")):
                    self.log("[INFO] Security check passed.")
                    return True
            self.log("[ERROR] Security check timed out.")
            return False

        self.log("[ERROR] Login failed.")
        return False

    # ─────────────────────────────────── post extraction

    def _get_posts(self) -> list:
        time.sleep(1.5)
        for sel in POST_CONTAINER_SELECTORS:
            posts = self.driver.find_elements(By.CSS_SELECTOR, sel)
            visible = [p for p in posts if p.is_displayed()]
            if len(visible) >= 1:
                return visible
        return []

    def _get_urn(self, post) -> str:
        """Unique identifier for a post — used to avoid saving duplicates."""
        for attr in ("data-urn", "data-id"):
            val = post.get_attribute(attr) or ""
            if "activity" in val:
                return val
        return ""

    def _get_post_url(self, post) -> str:
        """
        Build a canonical post URL.
        LinkedIn feed post permalinks look like:
        https://www.linkedin.com/feed/update/urn:li:activity:1234567890/
        """
        urn = self._get_urn(post)
        if urn:
            # Normalise urn:li:ugcPost:... → urn:li:activity:...
            # (both redirect correctly on LinkedIn)
            return f"https://www.linkedin.com/feed/update/{urn}/"

        # Fallback: find any feed/update link inside the post
        try:
            for a in post.find_elements(By.TAG_NAME, "a"):
                href = a.get_attribute("href") or ""
                if "feed/update" in href or "ugcPost" in href:
                    return href.split("?")[0]
        except Exception:
            pass
        return ""

    def _get_author_name(self, post) -> str:
        for sel in AUTHOR_NAME_SELECTORS:
            try:
                t = post.find_element(By.CSS_SELECTOR, sel).text.strip()
                if t:
                    return t
            except NoSuchElementException:
                continue
        return "Unknown Author"

    def _get_author_url(self, post) -> str:
        for sel in AUTHOR_URL_SELECTORS:
            try:
                href = post.find_element(By.CSS_SELECTOR, sel).get_attribute("href") or ""
                if href:
                    return href.split("?")[0]
            except NoSuchElementException:
                continue
        return ""

    def _get_post_text(self, post) -> str:
        for sel in POST_TEXT_SELECTORS:
            try:
                t = post.find_element(By.CSS_SELECTOR, sel).text.strip()
                if len(t) > 20:
                    return t
            except NoSuchElementException:
                continue
        return ""

    def _keywords_in_text(self, text: str) -> list[str]:
        lower = text.lower()
        return [
            kw for kw in self.config.get("feed_keywords", [])
            if kw.strip().lower() in lower
        ]

    # ─────────────────────────────────── search

    def _build_search_queries(self) -> list[str]:
        """
        Combine feed_keywords × job_roles to generate search queries.
        e.g. keywords=["C2C","Contract"], roles=["ai engineer"]
          → ["C2C ai engineer", "Contract ai engineer", "C2C", "Contract"]
        Deduplicates and caps at 20 queries.
        """
        keywords = [k.strip() for k in self.config.get("feed_keywords", []) if k.strip()]
        roles    = [r.strip() for r in self.config.get("job_roles", []) if r.strip()]

        queries: list[str] = []
        seen: set[str] = set()

        def _add(q: str):
            if q.lower() not in seen:
                seen.add(q.lower())
                queries.append(q)

        # keyword + role combinations first (most specific)
        for kw in keywords:
            for role in roles:
                _add(f"{kw} {role}")

        # then keyword alone
        for kw in keywords:
            _add(kw)

        return queries[:20]

    def _navigate_to_search(self, query: str):
        """Open LinkedIn post/content search for the given query."""
        encoded = query.replace(" ", "%20")
        url = (
            f"https://www.linkedin.com/search/results/content/"
            f"?keywords={encoded}&origin=GLOBAL_SEARCH_HEADER&sortBy=date_posted"
        )
        self.log(f"[INFO] Searching posts: '{query}'")
        self.driver.get(url)
        self._delay(3, 5)
        # Wait for results container
        try:
            WebDriverWait(self.driver, 12).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "div.search-results-container, div.scaffold-layout__list")
                )
            )
        except TimeoutException:
            self.log(f"[WARN] Search results slow to load for '{query}'")

    # ─────────────────────────────────── post helpers

    def _get_urn(self, post) -> str:
        for attr in ("data-urn", "data-id", "data-activity-urn"):
            val = post.get_attribute(attr) or ""
            if val:
                return val
        return ""

    def _get_post_url(self, post) -> str:
        urn = self._get_urn(post)
        if urn:
            return f"https://www.linkedin.com/feed/update/{urn}/"
        try:
            for a in post.find_elements(By.TAG_NAME, "a"):
                href = a.get_attribute("href") or ""
                if "feed/update" in href or "ugcPost" in href or "activity" in href:
                    return href.split("?")[0]
        except Exception:
            pass
        return self.driver.current_url.split("?")[0]

    def _get_author_name(self, post) -> str:
        for sel in AUTHOR_NAME_SELECTORS:
            try:
                t = post.find_element(By.CSS_SELECTOR, sel).text.strip()
                if t:
                    return t
            except NoSuchElementException:
                continue
        return "Unknown Author"

    def _get_author_url(self, post) -> str:
        for sel in AUTHOR_URL_SELECTORS:
            try:
                href = post.find_element(By.CSS_SELECTOR, sel).get_attribute("href") or ""
                if "/in/" in href:
                    return href.split("?")[0]
            except NoSuchElementException:
                continue
        return ""

    def _get_post_text(self, post) -> str:
        # Try dedicated text selectors
        for sel in POST_TEXT_SELECTORS:
            try:
                t = post.find_element(By.CSS_SELECTOR, sel).text.strip()
                if len(t) > 20:
                    return t
            except NoSuchElementException:
                continue
        # Fallback: all text inside the card
        try:
            t = post.text.strip()
            if len(t) > 20:
                return t
        except Exception:
            pass
        return ""

    def _keywords_in_text(self, text: str) -> list[str]:
        lower = text.lower()
        return [kw for kw in self.config.get("feed_keywords", [])
                if kw.strip().lower() in lower]

    def _process_post(self, post, found_count: int) -> int:
        """Extract data from one post element; save if keywords match. Returns new found_count."""
        try:
            urn = self._get_urn(post)
            if urn and urn in self._seen_urns:
                return found_count
            if urn:
                self._seen_urns.add(urn)

            text    = self._get_post_text(post)
            matched = self._keywords_in_text(text) if text else []
            if not matched:
                return found_count

            author     = self._get_author_name(post)
            author_url = self._get_author_url(post)
            post_url   = self._get_post_url(post)
            now        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            preview    = text[:250].replace("\n", " ")
            kw_str     = ", ".join(matched)

            found_count += 1
            self.found_posts.append({
                "author":        author,
                "author_url":    author_url,
                "post_url":      post_url,
                "preview":       preview,
                "keywords_found": kw_str,
                "timestamp":     now,
            })
            if self.count_callback:
                self.count_callback(found_count)

            self.log(
                f"[FOUND] #{found_count}  {author}  |  {kw_str}\n"
                f"        {post_url}\n"
                f"        {preview[:120]}..."
            )

        except StaleElementReferenceException:
            pass
        except Exception as exc:
            self.log(f"[WARN] Post parse error: {exc}")

        return found_count

    # ─────────────────────────────────── main run

    def run(self):
        try:
            self._setup_driver()
            if not self._login():
                return

            queries     = self._build_search_queries()
            max_scrolls = int(self.config.get("feed_max_scrolls", 30))
            found_count = 0

            # Distribute scrolls evenly across queries
            scrolls_per_query = max(3, max_scrolls // max(len(queries), 1))

            self.log(f"[INFO] {len(queries)} search queries | {scrolls_per_query} scrolls each")
            self.log(f"[INFO] Queries: {queries}")

            for query in queries:
                if self._should_stop():
                    break

                self._navigate_to_search(query)

                for scroll_num in range(scrolls_per_query):
                    if self._should_stop():
                        break

                    posts = self._get_posts()
                    self.log(f"[DEBUG] '{query}' scroll {scroll_num+1} — {len(posts)} items")

                    for post in posts:
                        if self._should_stop():
                            break
                        found_count = self._process_post(post, found_count)

                    self.driver.execute_script("window.scrollBy(0, window.innerHeight * 1.8);")
                    self._delay(2.5, 4.5)

            self.log(f"\n[DONE] Scan complete — {found_count} posts found across {len(queries)} searches")
            update_feed_excel(self.found_posts, self.log)

        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass
                self.log("[INFO] Feed scanner browser closed.")
